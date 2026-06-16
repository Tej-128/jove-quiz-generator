"""
JoVE Quiz Generator - Excel Export v3.1
Writes Summary first, then Set 1 / Set 2 / Set 3 sheets with the exact required columns.
"""

from __future__ import annotations

from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

QUESTION_TYPES = [
    "Single Correct",
    "Multi Correct",
    "True or False",
    "Fill in the Blanks",
    "Dropdown",
    "Match the following",
    "Categorisation",
]

MIN_PER_TYPE = 5
QUESTIONS_PER_SET = 40
NUM_SETS = 3

COLUMNS = [
    ("Chapter Name", 25),
    ("Video ID", 12),
    ("Question Index", 14),
    ("Question Content", 55),
    ("Question Type", 18),
    ("Option 1", 45),
    ("Option 2", 45),
    ("Option 3", 45),
    ("Option 4", 45),
    ("Right Answer", 15),
]

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
SUMMARY_FILL = PatternFill("solid", fgColor="ED7D31")
SUMMARY_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
BODY_BOLD = Font(name="Arial", bold=True, size=10)
OK_FILL = PatternFill("solid", fgColor="E2EFDA")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
FAIL_FILL = PatternFill("solid", fgColor="F4CCCC")
SET_FILLS = [
    PatternFill("solid", fgColor="DEEAF1"),
    PatternFill("solid", fgColor="E2EFDA"),
    PatternFill("solid", fgColor="FFF2CC"),
]
THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _cell(ws, row: int, col: int, value: Any, font=None, fill=None, align=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    if border:
        cell.border = border
    return cell


def _write_headers(ws, row: int = 1) -> None:
    for col_idx, (name, width) in enumerate(COLUMNS, 1):
        _cell(
            ws,
            row,
            col_idx,
            name,
            font=HEADER_FONT,
            fill=HEADER_FILL,
            align=Alignment(horizontal="center", vertical="center", wrap_text=True),
            border=THIN,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[row].height = 30


def _right_answer_value(question: dict[str, Any]):
    question_type = question.get("question_type", "")
    if question_type in {"Match the following", "Categorisation"}:
        return None
    value = question.get("right_answer", "")
    if value is None:
        return None
    value = str(value)
    if value.lower() == "none":
        return None
    return value


def _write_question_row(ws, row: int, question: dict[str, Any], chapter_name: str, fill=None) -> None:
    values = [
        chapter_name,
        str(question.get("lesson_id", "")),
        question.get("question_index", ""),
        question.get("question_content", ""),
        question.get("question_type", ""),
        question.get("option_1", "") or None,
        question.get("option_2", "") or None,
        question.get("option_3", "") or None,
        question.get("option_4", "") or None,
        _right_answer_value(question),
    ]
    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font = BODY_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = THIN
        if fill:
            cell.fill = fill
    ws.row_dimensions[row].height = 55


def _set_status(question_type: str, count: int, set_report: dict[str, Any]) -> str:
    budget = set_report.get("budget", {}) or {}
    dropped = set_report.get("dropped_types", []) or []
    if question_type in dropped or question_type not in budget:
        return "Not generated"
    required = int(budget.get(question_type, MIN_PER_TYPE))
    if count == 0:
        return "Not generated"
    if count < MIN_PER_TYPE or count < required:
        return "Below minimum"
    return "OK"


def _status_fill(status: str):
    if status == "OK":
        return OK_FILL
    if status == "Below minimum":
        return WARN_FILL
    return FAIL_FILL


def _write_summary(wb: openpyxl.Workbook, all_sets: list[list[dict[str, Any]]], report: dict[str, Any], chapter_name: str) -> None:
    ws = wb.active
    ws.title = "Summary"
    for col_idx, width in enumerate([28, 16, 16, 16, 16, 16, 16, 16, 16, 55], 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    resolved_chapter = report.get("chapter_name") or chapter_name
    total_questions = sum(len(question_set) for question_set in all_sets)

    row = 1
    title = f"Quiz Generation Summary - {resolved_chapter}"
    _cell(
        ws,
        row,
        1,
        title,
        font=Font(name="Arial", bold=True, size=13, color="FFFFFF"),
        fill=SUMMARY_FILL,
        align=Alignment(horizontal="center", vertical="center"),
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    ws.row_dimensions[row].height = 28
    row += 2

    info_rows = [
        ("Chapter Name", resolved_chapter),
        ("Number of Lessons", report.get("num_lessons", 0)),
        ("Content Lessons Used", report.get("num_content_lessons", 0)),
        ("Total Questions", total_questions),
        ("Expected Questions", NUM_SETS * QUESTIONS_PER_SET),
        ("Sets Generated", len(all_sets)),
        ("Skipped Empty Lessons", ", ".join(report.get("skipped_lessons", []) or []) or "None"),
        ("Dropped Types", ", ".join(report.get("fallback_types_dropped", []) or []) or "None"),
    ]
    for label, value in info_rows:
        _cell(ws, row, 1, label, font=BODY_BOLD)
        _cell(ws, row, 2, value, font=BODY_FONT)
        row += 1
    row += 1

    headers = [
        "Question Type",
        "Set 1 Count",
        "Set 1 Status",
        "Set 2 Count",
        "Set 2 Status",
        "Set 3 Count",
        "Set 3 Status",
        "Total",
        "Minimum / Set",
        "Notes",
    ]
    for col_idx, header in enumerate(headers, 1):
        _cell(
            ws,
            row,
            col_idx,
            header,
            font=SUMMARY_FONT,
            fill=SUMMARY_FILL,
            align=Alignment(horizontal="center", vertical="center", wrap_text=True),
            border=THIN,
        )
    ws.row_dimensions[row].height = 28
    row += 1

    set_reports = report.get("sets", []) or []
    for question_type in QUESTION_TYPES:
        row_values: list[Any] = [question_type]
        total_for_type = 0
        statuses: list[str] = []
        notes: list[str] = []
        for set_index in range(NUM_SETS):
            set_report = set_reports[set_index] if set_index < len(set_reports) else {}
            count = int((set_report.get("type_counts", {}) or {}).get(question_type, 0))
            status = _set_status(question_type, count, set_report)
            total_for_type += count
            statuses.append(status)
            row_values.extend([count, status])
            if status != "OK":
                notes.append(f"Set {set_index + 1}: {status}")
        row_values.extend([total_for_type, MIN_PER_TYPE, "; ".join(notes)])

        for col_idx, value in enumerate(row_values, 1):
            fill = None
            if col_idx in {3, 5, 7}:
                fill = _status_fill(str(value))
            _cell(
                ws,
                row,
                col_idx,
                value,
                font=BODY_FONT,
                fill=fill,
                align=Alignment(horizontal="center" if col_idx > 1 else "left", wrap_text=True),
                border=THIN,
            )
        row += 1

    row += 2

    skipped_files = report.get("skipped_files", []) or []
    if skipped_files:
        _cell(ws, row, 1, "Skipped / Ignored Files", font=BODY_BOLD, fill=WARN_FILL)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        row += 1
        for skipped in skipped_files:
            text = f"{skipped.get('name', '')}: {skipped.get('reason', '')}"
            _cell(ws, row, 1, text, font=BODY_FONT, fill=WARN_FILL, align=Alignment(wrap_text=True))
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
            row += 1
        row += 1

    warnings = report.get("warnings", []) or []
    if warnings:
        _cell(ws, row, 1, f"Warnings / Flags ({len(warnings)})", font=BODY_BOLD, fill=WARN_FILL)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        row += 1
        for warning in warnings:
            _cell(ws, row, 1, warning, font=BODY_FONT, fill=WARN_FILL, align=Alignment(wrap_text=True))
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
            row += 1
    else:
        _cell(ws, row, 1, "No warnings or flags.", font=BODY_BOLD, fill=OK_FILL)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)

    ws.freeze_panes = "A12"


def build_excel(all_sets: list[list[dict[str, Any]]], report: dict[str, Any], chapter_name: str) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    resolved_chapter = report.get("chapter_name") or chapter_name
    _write_summary(wb, all_sets, report, resolved_chapter)

    for set_index, questions in enumerate(all_sets, 1):
        ws = wb.create_sheet(title=f"Set {set_index}")
        ws.freeze_panes = "A2"
        _write_headers(ws)
        fill = SET_FILLS[(set_index - 1) % len(SET_FILLS)]
        for row_index, question in enumerate(questions, 2):
            row_fill = fill if row_index % 2 == 0 else None
            _write_question_row(ws, row_index, question, resolved_chapter, fill=row_fill)

    return wb


def save_workbook(wb: openpyxl.Workbook, output_path: str) -> None:
    wb.save(output_path)
